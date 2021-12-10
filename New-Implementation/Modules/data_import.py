# Imports packages
import openpyxl
from typing import Dict, Tuple, List, Callable
from dataclasses import dataclass

from Modules.data_classes import ppe_data_class, model_param_class, transition_data_class, input_data_class

# Reads All input data from excel
def read_data(data_file_path):

    #Opens the excel book
    book = openpyxl.load_workbook(data_file_path, data_only=True)


    # Generates Indices
    indices_sheet = book.get_sheet_by_name('Indices Data')

    # Generates T
    t = []
    for t_row in indices_sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        if t_row[0] == None: break
        t.append(t_row[0])
    # Generates M
    m = []
    for m_row in indices_sheet.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
        if m_row[0] == None: break
        m.append(m_row[0])
    # Generates P
    p = []
    for ppe_row in indices_sheet.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True):
        if ppe_row[0] == None: break
        p.append(ppe_row[0])
    # Generates D
    d = []
    for comp_row in indices_sheet.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True):
        if comp_row[0] == None: break
        d.append(comp_row[0])
    # Generates K
    k = []
    for prio_row in indices_sheet.iter_rows(min_row=2, min_col=5, max_col=5, values_only=True):
        if prio_row[0] == None: break
        k.append(prio_row[0])
    # Generates C
    c = []
    for cpu_row in indices_sheet.iter_rows(min_row=2, min_col=6, max_col=6, values_only=True):
        if cpu_row[0] == None: break
        c.append(cpu_row[0])

    # Saves Indices
    indices = {'t': t, 'm': m, 'p': p, 'd': d, 'k': k, 'c': c }
    

    # PPE Data
    ppe_data_sheet = book.get_sheet_by_name('PPE Data')

    ppe_data = {}
    for ppe_row in ppe_data_sheet.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True):
        if ppe_row[0] == None: break
        ppe_type = ""
        if ppe_row[1] == True: ppe_type = "carry-over"
        else: ppe_type = "non-carry-over"
        ppe_data[ppe_row[0]] = ppe_data_class(ppe_type, ppe_row[2], [int(i) for i in ppe_row[3].split(',')])


    # PPE Usage
    usage_sheet = book.get_sheet_by_name('PPE Usage')

    usage = {}
    for usage_row in usage_sheet.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True):
        if usage_row[0] == None: break
        usage[
            (usage_row[2], usage_row[1], usage_row[0])
        ] = usage_row[3]


    # Arrival Rates
    arrival_sheet = book.get_sheet_by_name('Patient Arrival')
    arrival = {}
    for arr_row in arrival_sheet.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True):
        if arr_row[0] == None: break

        arrival[
            (arr_row[1], arr_row[2], arr_row[0])
        ] = arr_row[3]


    # Patient Transitions
    transition_sheet = book.get_sheet_by_name('Patient Transitions')
    wait_limit = {}
    transition_comp = {}
    transition_pri = {}
    for row in transition_sheet.iter_rows(min_row=2, min_col=1, max_col=2, values_only=True):
        if row[0] == None: break
        wait_limit[row[0]] = row[1]
    for row in transition_sheet.iter_rows(min_row=2, min_col=3, max_col=5, values_only=True):
        if row[0] == None: break
        transition_comp[(row[0], row[1])] = row[2]
    for row in transition_sheet.iter_rows(min_row=2, min_col=6, max_col=8, values_only=True):
        if row[0] == None: break
        transition_pri[(row[0], row[1])] = row[2]
    transition = transition_data_class(wait_limit, transition_comp, transition_pri)

    # Model Parameters
    model_param_sheet = book.get_sheet_by_name('Model Parameters')
    cw = {}
    cc = {}
    cs = {}
    for row in model_param_sheet.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):
        if row[0] == None: break
        cw[row[0]] = row[1]
        cc[row[0]] = row[2]
        cs[row[0]] = [0]
    
    cv = model_param_sheet['D2'].value
    gamma = model_param_sheet['E2'].value

    # Cost of Scheduling
    for prior in k:
        for time in t:
            cs[prior].append(cs[prior][-1] + cw[prior] * (gamma**time) )
    
    model_param = model_param_class(cw, cc, cs, cv, gamma)

    # Expected Data
    expected_vals_sheet = book.get_sheet_by_name('Expected State Values')
    expected_vals = {}
    expected_vals['ul'] = {}
    for row in expected_vals_sheet.iter_rows(min_row=3, min_col=1, max_col=2, values_only=True):
        if row[0] == None: break
        expected_vals['ul'][row[0]] = row[1]
    expected_vals['pw'] = {}
    for row in expected_vals_sheet.iter_rows(min_row=3, min_col=3, max_col=7, values_only=True):
        if row[0] == None: break
        expected_vals['pw'][(row[0], row[1], row[2], row[3])] = row[4]
    expected_vals['ps'] = {}
    for row in expected_vals_sheet.iter_rows(min_row=3, min_col=8, max_col=13, values_only=True):
        if row[0] == None: break
        expected_vals['ps'][(row[0], row[1], row[2], row[3], row[4])] = row[5]


    # Returns data
    input_data = input_data_class(indices, ppe_data, usage, arrival, transition, model_param, expected_vals)
    return input_data
# Read betas from json
def read_betas(data_file_path):
    
    #Opens the excel book
    book = openpyxl.load_workbook(data_file_path, data_only=True)

    # Read Bets
    betas_sheet = book.get_sheet_by_name('Betas')
    betas = {}

    betas['b0'] = {}
    betas['b0']['b_0'] = betas_sheet.cell(row=2, column=2).value

    betas['ul'] = {}
    for row in betas_sheet.iter_rows(min_row=2, min_col=3, max_col=4, values_only=True):
        if row[0] == None: break
        betas['ul'][(row[0],)] = row[1]

    betas['pw'] = {}
    for row in betas_sheet.iter_rows(min_row=2, min_col=5, max_col=9, values_only=True):
        if row[0] == None: break
        betas['pw'][(row[0], row[1], row[2], row[3])] = row[4]

    betas['ps'] = {}
    for row in betas_sheet.iter_rows(min_row=2, min_col=10, max_col=15, values_only=True):
        if row[0] == None: break
        betas['ps'][(row[0], row[1], row[2], row[3], row[4])] = row[5]

    return betas