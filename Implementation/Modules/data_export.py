# Imports packages
import openpyxl
from dataclasses import dataclass
from Modules.data_classes import state, action
from typing import Dict, Tuple, List, Callable

# Exports all state-action pairs to a workbook
def export_all_state_action(state_action_set, data_file_path):
    # Creates
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Prints
    count = 1
    for each in state_action_set:
        ws = wb.create_sheet(f"SA-{count}")
        export_state_action_pair(each, ws)
        count += 1

    # Saves
    wb.save(data_file_path)

# Exports a single state-action part to a worksheet
def export_state_action_pair(state_action, ws):
    st = state_action[0]
    ac = state_action[1]

    def export_single_set(sa_set, col, name):
        columns = 0
        count = 1

        for key, item in sa_set.items():
            columns = len(key)
            count +=1

            # Saves Key elements
            for column in range(columns):
                ws.cell(row = count, column = col+column).value = key[column]
            
            # Saves value
            ws.cell(row = count, column = col+columns).value = item

        # Names columns
        ws.cell(row=1, column = col).value = name
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+columns)       
        
        return (col+columns+1)
        
    column = 1
    column = export_single_set(st.ue_tp, column, 'State - Units Expected')
    column = export_single_set(st.uu_tp, column, 'State - Units Used')
    column = export_single_set(st.uv_tp, column, 'State - Units Violated')
    column = export_single_set(st.pw_mdc, column, 'State - Patients Waiting')
    column = export_single_set(st.ps_tmdc, column, 'State - Patients Scheduled Already')
    column = export_single_set(ac.sc_tmdc, column, 'Action - Patients Scheduled')
    column = export_single_set(ac.rsc_ttpmdc, column, 'Action - Patients Rescheduled')
    
