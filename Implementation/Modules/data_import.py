# Imports packages
import openpyxl
from dataclasses import dataclass
from typing import Dict, Tuple, List, Callable

@dataclass(frozen=True)
class ppe_data_class:
    ppe_type: str
    expected_units: int
    deviation: List[int]
@dataclass(frozen=True)
class model_param_class:
    cw: float
    cc: float
    M: float
    gamma: float
@dataclass(frozen=True)
class input_data_class:
    indices: Dict[ str, List[int] ]
    ppe_data: Dict[ str,ppe_data_class ]
    usage: Dict[ Tuple[str], List[int] ]
    arrival: Dict[ Tuple[str], List[int] ]
    transition: Dict[ Tuple[str], List[int] ]
    model_param: model_param_class
    expected_state_values: Dict[ str, Dict[Tuple[str], float] ]

# Reads All data
def read_data(data_file_path):
    
    #Opens the excel book
    book = openpyxl.load_workbook(data_file_path, data_only=True)


    # Generates Indices
    indices_sheet = book.get_sheet_by_name('Indices Data')

    # Generates T
    t = []
    for t_row in indices_sheet.iter_rows(min_row=2, min_col=1, max_col=1, values_only=True):
        if t_row[0] == None: break
        t.append(t_row[0])
    # Generates M
    m = [0]
    for m_row in indices_sheet.iter_rows(min_row=2, min_col=2, max_col=2, values_only=True):
        if m_row[0] == None: break
        m.append(m_row[0])
    # Generates P
    p = []
    for ppe_row in indices_sheet.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True):
        if ppe_row[0] == None: break
        p.append(ppe_row[0])
    # Generates D
    d = []
    for comp_row in indices_sheet.iter_rows(min_row=2, min_col=4, max_col=4, values_only=True):
        if comp_row[0] == None: break
        d.append(comp_row[0])
    # Generates C
    c = []
    for cpu_row in indices_sheet.iter_rows(min_row=2, min_col=5, max_col=5, values_only=True):
        if cpu_row[0] == None: break
        c.append(cpu_row[0])

    # Saves Indices
    indices = {'t': t, 'm': m, 'p': p, 'd': d, 'c': c }
    

    # PPE Data
    ppe_data_sheet = book.get_sheet_by_name('PPE Data')

    ppe_data = {}
    for ppe_row in ppe_data_sheet.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True):
        if ppe_row[0] == None: break
        ppe_type = ""
        if ppe_row[1] == True: ppe_type = "carry-over"
        else: ppe_type = "non-carry-over"
        ppe_data[ppe_row[0]] = ppe_data_class(ppe_type, ppe_row[2], [int(i) for i in ppe_row[3].split(',')])


    # PPE Usage
    usage_sheet = book.get_sheet_by_name('PPE Usage')

    usage = {}
    for usage_row in usage_sheet.iter_rows(min_row=2, min_col=1, max_col=4, values_only=True):
        if usage_row[0] == None: break
        usage[
            (usage_row[2], usage_row[1], usage_row[0])
        ] = usage_row[3]


    # Arrival Rates
    arrival_sheet = book.get_sheet_by_name('Patient Arrival')
    arrival = {}
    for arr_row in arrival_sheet.iter_rows(min_row=2, min_col=1, max_col=3, values_only=True):
        if arr_row[0] == None: break

        arrival[
            (arr_row[1], arr_row[0])
        ] = arr_row[2]


    # Patient Transitions
    transition_sheet = book.get_sheet_by_name('Patient Transitions')
    transition = {}
    for row in transition_sheet.iter_rows(min_row=2, min_col=2, max_col=5, values_only=True):
        if row[0] == None: break

        transition[
            (row[2], row[1], row[0])
        ] = row[3]


    # Model Parameters
    model_param_sheet = book.get_sheet_by_name('Model Parameters')
    model_param = model_param_class(
        model_param_sheet.cell(row=2, column=1).value,
        model_param_sheet.cell(row=2, column=2).value,
        model_param_sheet.cell(row=2, column=3).value,
        model_param_sheet.cell(row=2, column=4).value
    )

    # Expected State Values
    expected_state_sheet = book.get_sheet_by_name('Expected State Values')
    expected_state_values = {}
    # ue 
    expected_state_values['ue'] = {}
    for row in expected_state_sheet.iter_rows(min_row=3, min_col=1,max_col=3, values_only=True):
        if row[0] == None: break
        key = (row[1], row[0])
        value = float(row[2])
        expected_state_values['ue'][key] = value
    # uu
    expected_state_values['uu'] = {}
    for row in expected_state_sheet.iter_rows(min_row=3, min_col=4,max_col=6, values_only=True):
        if row[0] == None: break
        key = (row[1], row[0])
        value = float(row[2])
        expected_state_values['uu'][key] = value
    # uv
    expected_state_values['uv'] = {}
    for row in expected_state_sheet.iter_rows(min_row=3, min_col=7,max_col=9, values_only=True):
        if row[0] == None: break
        key = (row[1], row[0])
        value = float(row[2])
        expected_state_values['uv'][key] = value
    # pe
    expected_state_values['pe'] = {}
    for row in expected_state_sheet.iter_rows(min_row=3, min_col=10,max_col=12, values_only=True):
        if row[0] == None: break
        key = (row[0], row[1])
        value = float(row[2])
        expected_state_values['pe'][key] = value
    # pw
    expected_state_values['pw'] = {}
    for row in expected_state_sheet.iter_rows(min_row=3, min_col=13,max_col=16, values_only=True):
        if row[0] == None: break
        key = (row[0], row[1], row[2])
        value = float(row[3])
        expected_state_values['pw'][key] = value
    # ps
    expected_state_values['ps'] = {}
    for row in expected_state_sheet.iter_rows(min_row=3, min_col=17,max_col=21, values_only=True):
        if row[0] == None: break
        key = (row[0], row[1], row[2], row[3])
        value = float(row[4])
        expected_state_values['ps'][key] = value

    # Returns data
    input_data = input_data_class(indices, ppe_data, usage, arrival, transition, model_param, expected_state_values)
    return input_data