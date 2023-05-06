# Imports packages
import openpyxl
from dataclasses import dataclass
from typing import Dict, Tuple, List, Callable

from Modules.data_classes import state, action

# Exports all state-action pairs to a workbook
def export_all_state_action(state_action_set, data_file_path):
    # Creates
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Prints
    count = 1
    for each in state_action_set:
        ws = wb.create_sheet(f"SA-{count}")
        export_state_action_pair(each, ws)
        count += 1

    # Saves
    wb.save(data_file_path)
# Exports a single state-action part to a worksheet
def export_state_action_pair(state_action, ws):
    st = state_action[0]
    ac = state_action[1]

    def export_single_set(sa_set, col, name):
        columns = 0
        count = 1

        for key, item in sa_set.items():
            columns = len(key)
            count +=1

            # Saves Key elements
            for column in range(columns):
                ws.cell(row = count, column = col+column).value = key[column]
            
            # Saves value
            ws.cell(row = count, column = col+columns).value = item

        # Names columns
        ws.cell(row=1, column = col).value = name
        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col+columns)       
        
        return (col+columns+1)
        
    column = 1
    column = export_single_set(st.ue_tp, column, 'State - Units Expected')
    column = export_single_set(st.uu_tp, column, 'State - Units Used')
    column = export_single_set(st.pw_mdc, column, 'State - Patients Waiting')
    column = export_single_set(st.ps_tmdc, column, 'State - Patients Scheduled Already')

    column = export_single_set(ac.sc_tmdc, column, 'Action - Patients Scheduled')
    column = export_single_set(ac.rsc_ttpmdc, column, 'Action - Patients Rescheduled')
    column = export_single_set(ac.uv_tp, column, 'Action - Units Violated')

    column = export_single_set(ac.uu_p_tp, column, 'Action - Units Used - Post Action')
    column = export_single_set(ac.pw_p_mdc, column, 'Action - Patients Waiting - Post Action')
    column = export_single_set(ac.ps_p_tmdc, column, 'Action - Patients Scheduled Already - Post Action')
# Export Betas to Excel
def export_betas(betas, data_file_path):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet(f"Betas")

    # Prints all
    col_count = 1
    # Iterates over states
    for state in betas.keys():
        row_count = 2
        col_length = 0
        ws.cell(row=1, column=col_count).value = state

        # Iterates within state
        for key in betas[state].keys():

            if type(key) is tuple:
                col_length = len(key)
                for index in range(len(key)):
                    ws.cell(row=row_count, column=col_count+index).value = key[index]
            else: 
                col_length = 1
                ws.cell(row=row_count, column=col_count).value = key

            ws.cell(row=row_count, column=col_count+col_length).value = betas[state][key]
            row_count += 1
        
        col_count += col_length + 1 

    # Saves
    wb.save(data_file_path)
# Export Expected state Values
def export_expected_vals(data_file_path, state):
    wb = openpyxl.load_workbook(data_file_path, data_only=True)
    ws = wb["Expected State Values"]

    # UE
    count = 3
    for tp, v in state.ue_tp.items():
        ws.cell(row=count, column=1).value = tp[1]
        ws.cell(row=count, column=2).value = tp[0]
        ws.cell(row=count, column=3).value = v
        count += 1

    # UU
    count = 3
    for tp, v in state.uu_tp.items():
        ws.cell(row=count, column=4).value = tp[1]
        ws.cell(row=count, column=5).value = tp[0]
        ws.cell(row=count, column=6).value = v
        count += 1

    # PW
    count = 3
    for mdc, v in state.pw_mdc.items():
        ws.cell(row=count, column=7).value = mdc[0]
        ws.cell(row=count, column=8).value = mdc[1]
        ws.cell(row=count, column=9).value = mdc[2]
        ws.cell(row=count, column=10).value = v
        count += 1

    # UE
    count = 3
    for tmdc, v in state.ps_tmdc.items():
        ws.cell(row=count, column=11).value = tmdc[0]
        ws.cell(row=count, column=12).value = tmdc[1]
        ws.cell(row=count, column=13).value = tmdc[2]
        ws.cell(row=count, column=14).value = tmdc[3]
        ws.cell(row=count, column=15).value = v
        count += 1
        
    wb.save(data_file_path)